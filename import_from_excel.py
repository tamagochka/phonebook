from openpyxl import load_workbook

import mysql.connector
import sys
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', default='phones.xlsx', type=str, help='Путь до файла Excel.')
parser.add_argument('-s', '--server', default='localhost', type=str, help='Сервер базы данных.')
parser.add_argument('-u', '--user', default='root', type=str, help='Пользователь базы данных.')
parser.add_argument('-p', '--password', default='password', type=str, help='Пароль от базы данных.')
parser.add_argument('-b', '--base', default='phonebook', type=str, help='Имя базы данных.')
args = parser.parse_args(sys.argv[1:])

# 'C:/Users/tamagochka/OneDrive/Рабочий стол/тлф.xlsx'
path_to_excel_file = args.file
host_db = args.server
user_db = args.user
pass_db = args.password
name_db = args.base

db = mysql.connector.connect(host=host_db, user=user_db, password=pass_db, database=name_db)
cursor = db.cursor()

wb = load_workbook(filename=path_to_excel_file)


wb.active = 1
ws = wb.active

answer = input('внести в базу названия рангов? [y/N]: ')
if answer.upper() == 'Y':
    insert_query = 'INSERT INTO ranks (`id`, `rank`) VALUES (%(id)s, %(rank)s)'
    for row in ws.iter_rows():
        rank = {
            'id': row[0].value,
            'rank': row[1].value if row[1].value else ''
        }
        cursor.execute(insert_query, rank)
    db.commit()


wb.active = 0
ws = wb.active

answer = input('внести в базу названия подразделений? [y/N]: ')
if answer.upper() == 'Y':
    parent_id = 0
    insert_query = 'INSERT INTO departments (`title`, `parent_id`) VALUES (%(title)s, %(parent_id)s)'
    get_id_query = 'SELECT id FROM departments WHERE title=%s'
    # заполняем таблицу с названиями подразелений
    for row in ws.iter_rows():
        # for row in ws['358:413']:
        # если есть данные только в первой ячейке, то это название корневого подразделения
        if row[0].value and not row[1].value:
            department = {
                'title': row[0].value,
                'parent_id': 0  # корневое подразделение
            }
            cursor.execute(insert_query, department)
            # получаем id под которым подразделения было добавлено в базу
            # если после него будет идти дочернее подразделения, то этот id будет добавлен ему в parent_id
            cursor.execute(get_id_query, (row[0].value,))
            r = cursor.fetchone()
            parent_id = r[0]
            print(parent_id, end=' ')
            print(row[0].value)
        # если данные есть только во второй ячейке, то это дочернее подразделение
        if not row[0].value and row[1].value:
            department = {
                'title': row[1].value,
                'parent_id': parent_id  # дочернему подразделению добавляем id предыдущего
            }
            cursor.execute(insert_query, department)
            print('\t' + row[1].value)
    db.commit()

# проверяем названия рангов, записанных в таблице excel,
# чтобы они соответствовали названиям рангов, записанных
# в таблице ranks базы данных
answer = input('проверить название рангов в файле на соответствие базе? [y/N]: ')
if answer.upper() == 'Y':
    query = 'SELECT `rank` FROM ranks'
    cursor.execute(query)
    ranks = cursor.fetchall()
    ranks = [rank[0] for rank in ranks]
    n = 0
    for row in ws.iter_rows():
        n += 1
        if row[2].value in ranks or not row[2].value:
            continue
        else:
            print('ошибка в ранге в строке:', n, '-', row[2].value)

# проверка на наличие ошибок в ФИО, ячейка с ФИО должна содержать 3 слова
answer = input('проверить наличие ошибок в ФИО? [y/N]: ')
if answer.upper() == 'Y':
    n = 0
    for row in ws.iter_rows():
        n += 1
        if row[0].value and row[1].value:
            try:
                (surname, name, patronymic) = row[1].value.split()
            except ValueError:
                print('ошибка в ФИО в строке:', n, '-', row[1].value)

# внести в базу абонентов
answer = input('внести в базу данных абонентов? [y/N]: ')
if answer.upper() == 'Y':
    get_rank_id_query = 'SELECT `id` FROM ranks WHERE rank=%s'
    get_department_id_query = 'SELECT `id` FROM departments WHERE title=%s'
    insert_abonent_query = 'INSERT INTO abonents' \
        '(`post`, `surname`, `name`, `patronymic`, `rank`, `extension_number`, `landline_number`, `department`) VALUES'\
        '(%(post)s, %(surname)s, %(name)s, %(patronymic)s, %(rank)s,' \
        '%(extension_number)s, %(landline_number)s, %(department)s)'
    department_id = 0
    for row in ws.iter_rows():
        # если строка с названием подразделения, то получаем его id
        if row[0].value and not row[1].value:
            cursor.execute(get_department_id_query, (row[0].value,))
            department_id = cursor.fetchone()[0]
        if not row[0].value and row[1].value:
            cursor.execute(get_department_id_query, (row[1].value,))
            department_id = cursor.fetchone()[0]
        # если строка с абонентом, то заносим его данные в базу
        if row[0].value and row[1].value:
            rank = row[2].value if row[2].value else ''
            cursor.execute(get_rank_id_query, (rank,))
            rank_id = cursor.fetchone()[0]
            (surname, name, patronymic) = row[1].value.split()
            abonent = {
                'post': row[0].value,
                'surname': surname,
                'name': name,
                'patronymic': patronymic,
                'rank': rank_id,
                'extension_number': row[3].value if row[3].value else '',
                'landline_number': row[4].value if row[4].value else '',
                'department': department_id
            }
            cursor.execute(insert_abonent_query, abonent)
    db.commit()


wb.close()
cursor.close()
db.close()
